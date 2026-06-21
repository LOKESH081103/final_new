"""
DHC Working automation — core ETL pipeline.

Replicates, in pandas, the manual VLOOKUP/COUNTIFS/pivot workflow found inside
DHC_Working_-_Jun_26.xlsx. See ROADMAP.md for the full reverse-engineered logic
map and the open questions that need confirming with the process owner.
"""
import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Static mapping tables copied verbatim from the workbook's "Sheet3" tab.
# These are the small reference tables mam's VLOOKUPs point at.
# ---------------------------------------------------------------------------
MODE_MAP = {
    "AIRTEL": "AIRTEL / CASH", "CASH": "AIRTEL / CASH",
    "CHEQUE": "CHQ / DD", "DD": "CHQ / DD",
    "ONLINE_PAYMENT": "ONLINE_PAYMENT", "RTGS": "RTGS",
}
STATUS_MAP = {"B": "Bounced", "C": "Cleared", "D": "Deposit", "NA": "Pending", "X": "Cxn"}
RECEIPT_SOURCE_MAP = {
    "BBPS": "BBPS", "CHOLAONE DIRECT": "CHOLAONE DIRECT",
    "CCP - BITLY": "CCP - BITLY", "CCP - QR": "CCP - QR", "CCP": "CCP - QR",
}
RECEIPT_TYPE_MAP = {
    "Part Payment": "Part Payment", "FC": "Settlement",
    "Sale/EMD receipt": "Settlement", "Settlement": "Settlement", "OD": "OD",
}
RECEIPT_TYPE_FALLBACK = {"OD": "OD", "OTHER OD": "OTHER OD"}


EXCEL_DATE_COLUMNS = [
    "Date", "TXN DATE IN PL TAB", "RECEIPT ENTER DATE",
    "RECEIPTENTEREDTIME", "VALUEDATE", "TXNDATE",
]


def _fix_excel_dates(df: pd.DataFrame) -> pd.DataFrame:
    """
    pyxlsb (unlike openpyxl) returns Excel date cells as raw serial numbers,
    not datetimes. Convert the known date columns so downstream date-math
    behaves correctly.
    """
    for col in EXCEL_DATE_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], unit="D", origin="1899-12-30", errors="coerce")
    return df


def tat_bucket(days):
    """Replicates the approximate-match VLOOKUP against Sheet3!A:B."""
    if pd.isna(days):
        return None
    if days < 5:
        return "Less then 4"
    if days < 11:
        return "5 TO 10"
    return "Great then 10"


# ---------------------------------------------------------------------------
# Loaders — one per source file
# ---------------------------------------------------------------------------
def load_dcr(file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """DCR.xlsb -> (raw receipts [Sheet1], agreement slab master [Sheet2])."""
    receipts = pd.read_excel(file, sheet_name="Sheet1", engine="pyxlsb")
    receipts = _fix_excel_dates(receipts)
    master = pd.read_excel(file, sheet_name="Sheet2", engine="pyxlsb", usecols=[0, 1, 2, 3])
    master.columns = ["AGREEMENTNO", "OPENING_DPD", "OPNG_SLAB_TYPE", "OPENING_SLAB"]
    master = master.dropna(subset=["AGREEMENTNO"]).drop_duplicates("AGREEMENTNO", keep="last")
    return receipts, master


def load_disable_lists(file) -> tuple[pd.DataFrame, pd.DataFrame]:
    """To_be_Disabled.xlsb -> (cif_level_disable, agreement_level_disable)."""
    cif = pd.read_excel(file, sheet_name="CIF Level Disable", engine="pyxlsb", usecols=[0, 15])
    cif.columns = ["CIF_NO", "Status"]
    cif = cif.dropna(subset=["CIF_NO"])
    agr = pd.read_excel(file, sheet_name="Agreement Level Disble", engine="pyxlsb", usecols=[0, 5])
    agr.columns = ["AGREEMENTNO", "Status"]
    agr = agr.dropna(subset=["AGREEMENTNO"])
    return cif, agr


def load_employee_mobiles(file) -> set[str]:
    """CIFCL_CBSL_List.xlsx -> set of employee/agent mobile numbers (as strings)."""
    emp = pd.read_excel(file, sheet_name=0, usecols=["Mobile Number"])
    nums = emp["Mobile Number"].dropna().astype(str).str.replace(r"\.0$", "", regex=True)
    return set(nums)


def load_prior_lookup_master(file) -> pd.DataFrame:
    """
    Pull forward the cumulative AGREEMENTNO -> CIF_NO/Zone/Sub Region/Slab
    table from a previous DHC_Working.xlsx's 'Look Up' sheet (cols J:N).
    This is the carried-forward state — see ROADMAP.md Open Question #1.
    """
    df = pd.read_excel(file, sheet_name="Look Up", usecols="J:N", engine="openpyxl")
    df.columns = ["AGREEMENTNO", "CIF_NO", "ZONE_NEW", "SUB_REGION", "OPENING_SLAB"]
    return df.dropna(subset=["AGREEMENTNO"]).drop_duplicates("AGREEMENTNO", keep="last")


# ---------------------------------------------------------------------------
# Transform — build the working tabs
# ---------------------------------------------------------------------------
def build_lookup_master(prior_master: pd.DataFrame, dcr_master: pd.DataFrame) -> pd.DataFrame:
    """
    Refresh the cumulative agreement master: keep CIF_NO/Zone/Sub Region for
    agreements already known, refresh Opening Slab for everyone from this
    month's DCR Sheet2, and add new agreements with CIF/Zone/Sub Region
    left blank (flagged for manual completion — they aren't derivable from
    any of the 4 input files).
    """
    merged = dcr_master.merge(
        prior_master[["AGREEMENTNO", "CIF_NO", "ZONE_NEW", "SUB_REGION"]],
        on="AGREEMENTNO", how="left",
    )
    merged = merged.rename(columns={"OPNG_SLAB_TYPE": "OPENING_SLAB_LABEL"})
    merged["NEEDS_CIF_MAPPING"] = merged["CIF_NO"].isna()
    return merged[
        ["AGREEMENTNO", "CIF_NO", "ZONE_NEW", "SUB_REGION",
         "OPENING_SLAB_LABEL", "OPENING_SLAB", "NEEDS_CIF_MAPPING"]
    ]


def build_dcr_tab(
    receipts: pd.DataFrame,
    lookup_master: pd.DataFrame,
    agr_disable: pd.DataFrame,
    cif_disable: pd.DataFrame,
    employee_mobiles: set[str],
) -> pd.DataFrame:
    """Replicates DCR tab columns 74-84 (the 11 derived/lookup columns)."""
    df = receipts.copy()

    df = df.merge(
        lookup_master[["AGREEMENTNO", "CIF_NO", "OPENING_SLAB_LABEL", "NEEDS_CIF_MAPPING"]],
        on="AGREEMENTNO", how="left",
    )
    df["CIF"] = df["CIF_NO"]

    df["Unique Mob number"] = df.groupby("MOBILENO")["MOBILENO"].transform("count")
    df["Mob Num VS Emp Mob Num"] = df["MOBILENO"].astype(str).str.replace(
        r"\.0$", "", regex=True
    ).isin(employee_mobiles)

    df["Mode"] = df["MODEOFPAYMENT"].map(MODE_MAP)
    df["Status"] = df["STATUS IN TAB"].map(STATUS_MAP)
    df["Receipt Source"] = df["RECEIPTSOURCE"].map(RECEIPT_SOURCE_MAP)

    # Zone / Sub Region: the raw DCR extract already carries these per-row
    # (Sub Zone / SUB REGION columns) — no need for the CIF master here.
    # NB: the pivot reports group by the granular "Sub Zone" field
    # (EAST_1/EAST_2/NORTH_1.../SOUTH_1/SOUTH_2/WEST_1/WEST_2), not the
    # broad 4-way "ZONE NEW" field — confirmed against the original
    # workbook's row labels.
    df["Zone"] = df["ZONE NEW"]
    df["Sub Region"] = df["SUB REGION"]

    # Slab: prefer this month's fresh Sheet2 value (OPNG SLAB/SLAB already
    # in the raw extract); this matches the BT/BU formulas which re-pull
    # from DCR.xlsb's own Sheet2 rather than the stale Look Up master.
    df["Slab"] = df["SLAB"]

    df = df.merge(
        agr_disable.rename(columns={"Status": "Ag Level cash mode"}),
        on="AGREEMENTNO", how="left",
    )
    df = df.merge(
        cif_disable.rename(columns={"CIF_NO": "CIF", "Status": "CIF LEVEL"}),
        on="CIF", how="left",
    )
    return df


def build_rtgs_tab(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """Filters to RTGS-mode receipts and computes Ageing / TAT / Receipt Type."""
    rtgs = dcr_tab[dcr_tab["MODEOFPAYMENT"] == "RTGS"].copy()
    rtgs["Ageing"] = (
        pd.to_datetime(rtgs["RECEIPT ENTER DATE"]) - pd.to_datetime(rtgs["TXN DATE IN PL TAB"])
    ).dt.days
    rtgs["TAT"] = rtgs["Ageing"].apply(tat_bucket)
    rtgs["Receipt Type"] = rtgs["RECEIPTTYPE"].map(RECEIPT_TYPE_MAP)
    rtgs["Receipt Type"] = rtgs["Receipt Type"].fillna(
        rtgs["RECEIPT CAT"].map(RECEIPT_TYPE_FALLBACK)
    )
    return rtgs


# ---------------------------------------------------------------------------
# Display constants for the pivot-style summary sheets
# ---------------------------------------------------------------------------
RECEIPT_TYPE_ORDER = ["OD", "Settlement", "Part Payment", "OTHER OD"]
RECEIPT_TYPE_DISPLAY = {
    "OD": "EMI OD/Charges",
    "Settlement": "FORECLOSURE/SETTLEMENT",
    "Part Payment": "PART PAYMENT",
    "OTHER OD": "Other OD",
}
TAT_ORDER = ["Less then 4", "5 TO 10", "Great then 10"]
TAT_DISPLAY = {"Less then 4": "< 4 Days", "5 TO 10": "5 - 10 Days", "Great then 10": "> 10 Days"}
MODE_ORDER = ["AIRTEL / CASH", "CHQ / DD", "ONLINE_PAYMENT", "RTGS"]


# ---------------------------------------------------------------------------
# Summary sheet builders — each returns a plain-data structure (dicts/lists),
# kept separate from Excel writing so the numbers can be unit-tested or
# previewed in Streamlit without touching openpyxl.
# ---------------------------------------------------------------------------
def _tat_stats(df: pd.DataFrame, value_col: str = "AMOUNTPAID") -> dict:
    out = {"by_bucket": {}, "total_count": 0, "total_value": 0.0}
    for bucket in TAT_ORDER:
        sub = df[df["TAT"] == bucket]
        cnt = int(len(sub))
        val = float(sub[value_col].sum()) / 1e7
        out["by_bucket"][bucket] = {"count": cnt, "value": val}
        out["total_count"] += cnt
        out["total_value"] += val
    return out


def compute_zone_tat_matrix(df: pd.DataFrame, value_col: str = "AMOUNTPAID") -> dict:
    """
    Zone (Sub Zone) x Receipt Type x TAT bucket — the structure behind both
    'RTGS Summary' and 'Delay in RCPTING Summary'. Zone here means the
    granular 'Sub Zone' field (EAST_1, NORTH_2, SOUTH_1, ...), confirmed
    against the original workbook's row labels.
    """
    g = df.dropna(subset=["Sub Zone", "TAT"]).copy()
    zones = sorted(g["Sub Zone"].unique())
    blocks = []
    grand = _tat_stats(g.iloc[0:0], value_col)  # zeroed template
    for zone in zones:
        zdf = g[g["Sub Zone"] == zone]
        subtotal = _tat_stats(zdf, value_col)
        breakdown = [
            (code, RECEIPT_TYPE_DISPLAY[code], _tat_stats(zdf[zdf["Receipt Type"] == code], value_col))
            for code in RECEIPT_TYPE_ORDER
        ]
        blocks.append({"zone": zone, "subtotal": subtotal, "breakdown": breakdown})
        for bucket in TAT_ORDER:
            grand["by_bucket"][bucket]["count"] += subtotal["by_bucket"][bucket]["count"]
            grand["by_bucket"][bucket]["value"] += subtotal["by_bucket"][bucket]["value"]
        grand["total_count"] += subtotal["total_count"]
        grand["total_value"] += subtotal["total_value"]
    return {"zones": blocks, "grand_total": grand}


def compute_online_receipt_source_block(df: pd.DataFrame) -> dict:
    """Simple Receipt Source count among ONLINE_PAYMENT-mode rows."""
    sub = df[df["Mode"] == "ONLINE_PAYMENT"]
    counts = sub["RECEIPTSOURCE"].value_counts(dropna=True)
    rows = [(name, int(cnt)) for name, cnt in counts.items()]
    return {"rows": rows, "total": int(counts.sum())}


def build_rtgs_summary(rtgs_tab: pd.DataFrame, dcr_tab: pd.DataFrame) -> dict:
    """
    The Zone x Receipt Type x TAT matrix comes from the RTGS-filtered tab.
    The 'Online Payment — Receipt Source' mini-block is a secondary KPI on
    the same dashboard and is computed from the full month's DCR data (it
    would always be empty if computed from rtgs_tab, since Mode there is
    always 'RTGS').
    """
    return {
        "matrix": compute_zone_tat_matrix(rtgs_tab),
        "online_source_block": compute_online_receipt_source_block(dcr_tab),
    }


def build_delay_in_rcpting_summary(dcr_tab: pd.DataFrame) -> dict:
    full = dcr_tab.copy()
    full["Ageing"] = (
        pd.to_datetime(full["RECEIPT ENTER DATE"]) - pd.to_datetime(full["TXN DATE IN PL TAB"])
    ).dt.days
    full["TAT"] = full["Ageing"].apply(tat_bucket)
    full["Receipt Type"] = full["RECEIPTTYPE"].map(RECEIPT_TYPE_MAP)
    full["Receipt Type"] = full["Receipt Type"].fillna(
        full["RECEIPT CAT"].map(RECEIPT_TYPE_FALLBACK)
    )
    return {
        "matrix": compute_zone_tat_matrix(full),
        "online_source_block": compute_online_receipt_source_block(full),
    }


def _status_table(df: pd.DataFrame, status_groups: list[str], status_cols: list[str]) -> dict:
    """
    RECEIPT STATUS (row group) x Mode (row) x Status (column) counts,
    used for both halves of 'Receipt made summary'.
    """
    groups = []
    grand = {c: 0 for c in status_cols}
    grand_total = 0
    for status in status_groups:
        sdf = df[df["RECEIPT STATUS"] == status]
        if sdf.empty:
            continue
        modes = [m for m in MODE_ORDER if m in sdf["Mode"].unique()]
        rows = []
        group_totals = {c: 0 for c in status_cols}
        for mode in modes:
            mdf = sdf[sdf["Mode"] == mode]
            counts = {c: int((mdf["Status"] == c).sum()) for c in status_cols}
            row_total = sum(counts.values())
            rows.append({"mode": mode, "counts": counts, "total": row_total})
            for c in status_cols:
                group_totals[c] += counts[c]
                grand[c] += counts[c]
            grand_total += row_total
        groups.append({"status": status, "rows": rows, "totals": group_totals})
    return {"groups": groups, "grand_totals": grand, "grand_total": grand_total}


def build_receipt_made_summary(dcr_tab: pd.DataFrame) -> dict:
    """
    Two side-by-side tables, both grouped from the same RECEIPT STATUS field:
    left = Updated/Pending x (Cleared, Deposited, Pending);
    right = Updated/Bounced-or-Cancelled x (Cleared, Deposited, Bounced, Cxn).
    """
    status_relabel = {
        "Updated": "UPDATED",
        "Updation Pending": "PENDING",
        "Bounced-or-Cancelled": "BOUNCED/CANCELLED",
    }
    d = dcr_tab.copy()
    d["RECEIPT STATUS"] = d["RECEIPT STATUS"].map(status_relabel).fillna(d["RECEIPT STATUS"])
    left = _status_table(d, ["UPDATED", "PENDING"], ["Cleared", "Deposit", "Pending"])
    right = _status_table(d, ["UPDATED", "BOUNCED/CANCELLED"], ["Cleared", "Deposit", "Bounced", "Cxn"])
    return {"left": left, "right": right}


def build_cash_mode_validation_summary(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Compliance check: customers flagged on either disable list who still
    paid in cash/Airtel, broken out by CIF and by day.
    """
    flagged = dcr_tab[
        (dcr_tab["Mode"] == "AIRTEL / CASH")
        & (dcr_tab["Ag Level cash mode"].notna() | dcr_tab["CIF LEVEL"].notna())
    ].copy()
    cols = ["CIF", "Zone2", "Sub Region2", "Slab", "Grand Total"]
    if flagged.empty:
        return pd.DataFrame(columns=cols)
    flagged["Receipt Date"] = pd.to_datetime(flagged["RECEIPT ENTER DATE"]).dt.normalize()
    pivot = pd.pivot_table(
        flagged, index=["CIF", "Zone", "Sub Region", "Slab"],
        columns="Receipt Date", values="AMOUNTPAID", aggfunc="sum",
    )
    pivot["Grand Total"] = pivot.sum(axis=1, skipna=True)
    pivot = pivot.reset_index().rename(columns={"Zone": "Zone2", "Sub Region": "Sub Region2"})
    return pivot


def build_rcpt_cxn(dcr_tab: pd.DataFrame) -> pd.DataFrame:
    """
    Cancelled-receipt register. Structural columns are auto-filled; the
    'Remarks' column is left blank for mam to fill in — this sheet is a
    judgment log, not a formula output (see ROADMAP.md).
    """
    cxn = dcr_tab[dcr_tab["Status"] == "Cxn"].copy()
    out = pd.DataFrame({
        "ReceiptNo": cxn.get("Receipt No"),
        "ReceiptDate": pd.to_datetime(cxn["RECEIPT ENTER DATE"]).dt.strftime("%d/%m/%Y"),
        "Amount": cxn["AMOUNTPAID"],
        "ReceiptStatus": "Cancelled",
        "ReceiptType": cxn.get("Receipt Type", cxn.get("RECEIPTTYPE")),
        "PaymentMode": cxn["Mode"],
        "Zone": cxn["Zone"],
        "AgreementNo": cxn["AGREEMENTNO"],
        "CustomerName": cxn.get("PAYERNAME"),
        "ReceiptCreatedDate": pd.to_datetime(cxn["RECEIPT ENTER DATE"]).dt.strftime("%d/%m/%Y"),
        "Status": "Duplicate Receipt",
        "Remarks": "",
    })
    return out


def _stats_row(label: str, stats: dict) -> list:
    """Helper to flatten a stats dict into a row for preview DataFrame."""
    row = [label]
    for bucket in TAT_ORDER:
        row += [stats["by_bucket"][bucket]["count"], round(stats["by_bucket"][bucket]["value"], 2)]
    row += [stats["total_count"], round(stats["total_value"], 2)]
    return row


def zone_tat_matrix_to_dataframe(matrix: dict) -> pd.DataFrame:
    """Flattens the Zone x Receipt Type x TAT structure into a preview-friendly DataFrame."""
    cols = ["Zone / Receipt Type"]
    for bucket in TAT_ORDER:
        cols += [f"{TAT_DISPLAY[bucket]} Count", f"{TAT_DISPLAY[bucket]} Value"]
    cols += ["Total Count", "Total Value (Cr)"]
    rows = []
    for block in matrix["zones"]:
        rows.append(_stats_row(block["zone"], block["subtotal"]))
        for code, label, stats in block["breakdown"]:
            rows.append(_stats_row(f"   {label}", stats))
    rows.append(_stats_row("GRAND TOTAL", matrix["grand_total"]))
    return pd.DataFrame(rows, columns=cols)


def receipt_made_table_to_dataframe(table: dict, status_cols: list[str]) -> pd.DataFrame:
    """Flattens the receipt status x mode table into a preview DataFrame."""
    cols = ["Status", "Mode"] + [c.upper() for c in status_cols] + ["Grand Total"]
    rows = []
    for group in table["groups"]:
        for i, r in enumerate(group["rows"]):
            label = group["status"] if i == 0 else ""
            rows.append([label, r["mode"]] + [r["counts"][c] for c in status_cols] + [r["total"]])
    rows.append(["GRAND TOTAL", ""] + [table["grand_totals"][c] for c in status_cols] + [table["grand_total"]])
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
SUBTOTAL_FILL = PatternFill("solid", start_color="BDD7EE")
GRANDTOTAL_FILL = PatternFill("solid", start_color="2E5395")
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color=NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=NAVY)
SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=10)
GRANDTOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="B7C5D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COUNT_FMT = '#,##0;-#,##0;"-"'
VALUE_FMT = '#,##0.00;-#,##0.00;"-"'
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_INDENT = Alignment(horizontal="left", indent=1)


def _set(ws, row, col, value, font=BODY_FONT, fill=None, fmt=None, align=None, border=BOX):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


def _write_df(ws, df: pd.DataFrame, start_row: int = 1):
    for j, col in enumerate(df.columns, start=1):
        col_label = col.strftime("%d-%b") if isinstance(col, pd.Timestamp) else str(col)
        _set(ws, start_row, j, col_label, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        width = max(10, min(28, len(col_label) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            fmt = None
            if isinstance(val, (np.integer,)):
                val = int(val)
                fmt = COUNT_FMT
            elif isinstance(val, (np.floating, float)):
                if pd.isna(val):
                    val = None
                else:
                    val = float(val)
                    fmt = VALUE_FMT
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
                fmt = "dd-mmm-yyyy"
            _set(ws, i, j, val, fmt=fmt)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _write_zone_tat_sheet(ws, title: str, summary: dict):
    """Writes the merged-header Zone x Receipt Type x TAT pivot, with an
    online-receipt-source mini block on the left, exactly mirroring the
    layout found in the original 'RTGS Summary' / 'Delay in RCPTING
    Summary' tabs."""
    matrix = summary["matrix"]
    source_block = summary["online_source_block"]

    ws.merge_cells("A1:L1")
    _set(ws, 1, 1, title, font=TITLE_FONT, fill=None, align=Alignment(horizontal="left"), border=None)

    # --- online receipt source mini-block (cols A:B) ---
    ws.merge_cells("A2:B2")
    _set(ws, 2, 1, "Online Payment — Receipt Source", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 3, 1, "Source", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 3, 2, "Receipt Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    r = 4
    for name, cnt in source_block["rows"]:
        _set(ws, r, 1, name, align=LEFT_INDENT)
        _set(ws, r, 2, cnt, fmt=COUNT_FMT)
        r += 1
    _set(ws, r, 1, "GRAND TOTAL", font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL)
    _set(ws, r, 2, source_block["total"], font=SUBTOTAL_FONT, fill=SUBTOTAL_FILL, fmt=COUNT_FMT)

    # --- main Zone x Receipt Type x TAT matrix (cols D:L) ---
    ws.merge_cells("D2:D3")
    _set(ws, 2, 4, "ZONE / RECEIPT TOWARDS", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    col = 5
    for bucket in TAT_ORDER:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        _set(ws, 2, col, TAT_DISPLAY[bucket], font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, col + 1, None, fill=HEADER_FILL)
        _set(ws, 3, col, "Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 3, col + 1, "Value (Cr)", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        col += 2
    ws.merge_cells("K2:K3")
    ws.merge_cells("L2:L3")
    _set(ws, 2, 11, "Total Count", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
    _set(ws, 2, 12, "Total Value (Cr)", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

    def _write_stat_row(row, label_col, label, stats, font, fill, indent=False):
        align = LEFT_INDENT if indent else Alignment(horizontal="left")
        _set(ws, row, label_col, label, font=font, fill=fill, align=align)
        col = label_col + 1
        for bucket in TAT_ORDER:
            _set(ws, row, col, stats["by_bucket"][bucket]["count"], font=font, fill=fill, fmt=COUNT_FMT)
            _set(ws, row, col + 1, stats["by_bucket"][bucket]["value"], font=font, fill=fill, fmt=VALUE_FMT)
            col += 2
        _set(ws, row, col, stats["total_count"], font=font, fill=fill, fmt=COUNT_FMT)
        _set(ws, row, col + 1, stats["total_value"], font=font, fill=fill, fmt=VALUE_FMT)

    row = 4
    for block in matrix["zones"]:
        _write_stat_row(row, 4, block["zone"], block["subtotal"], SUBTOTAL_FONT, SUBTOTAL_FILL)
        row += 1
        for code, display, stats in block["breakdown"]:
            _write_stat_row(row, 4, display, stats, BODY_FONT, None, indent=True)
            row += 1
    _write_stat_row(row, 4, "GRAND TOTAL", matrix["grand_total"], GRANDTOTAL_FONT, GRANDTOTAL_FILL)

    for letter, width in zip("ABCDEFGHIJKL", [20, 14, 3, 26, 9, 12, 9, 12, 9, 12, 12, 14]):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "E4"


def _write_receipt_made_summary_sheet(ws, summary: dict):
    def _write_table(start_col, title, status_cols, table):
        n = len(status_cols)
        last_col = start_col + 2 + n  # label cols (2) + status cols (n) + grand total (1)
        from openpyxl.utils import get_column_letter as gcl
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=last_col)
        _set(ws, 1, start_col, title, font=TITLE_FONT, align=Alignment(horizontal="left"), border=None)
        _set(ws, 2, start_col, "RECEIPT STATUS", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, start_col + 1, "PAYMENT MODE", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        for k, c in enumerate(status_cols):
            _set(ws, 2, start_col + 2 + k, c.upper(), font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        _set(ws, 2, last_col, "GRAND TOTAL", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)

        row = 3
        for group in table["groups"]:
            first_row = row
            for r in group["rows"]:
                _set(ws, row, start_col, None)
                _set(ws, row, start_col + 1, r["mode"], align=LEFT_INDENT)
                for k, c in enumerate(status_cols):
                    _set(ws, row, start_col + 2 + k, r["counts"][c], fmt=COUNT_FMT)
                _set(ws, row, last_col, r["total"], font=SUBTOTAL_FONT, fmt=COUNT_FMT)
                row += 1
            if row > first_row:
                _set(ws, first_row, start_col, group["status"], font=SUBTOTAL_FONT, align=Alignment(horizontal="left"))
                if row - 1 > first_row:
                    ws.merge_cells(start_row=first_row, start_column=start_col, end_row=row - 1, end_column=start_col)
        _set(ws, row, start_col, "GRAND TOTAL", font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL)
        _set(ws, row, start_col + 1, None, fill=GRANDTOTAL_FILL)
        for k, c in enumerate(status_cols):
            _set(ws, row, start_col + 2 + k, table["grand_totals"][c], font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
        _set(ws, row, last_col, table["grand_total"], font=GRANDTOTAL_FONT, fill=GRANDTOTAL_FILL, fmt=COUNT_FMT)
        return last_col

    last = _write_table(1, "Receipt Made Summary — Updated / Pending",
                         ["Cleared", "Deposit", "Pending"], summary["left"])
    _write_table(last + 2, "Receipt Made Summary — Updated / Bounced or Cancelled",
                 ["Cleared", "Deposit", "Bounced", "Cxn"], summary["right"])

    for letter, width in zip("ABCDEFGHIJKLMN", [20, 16, 11, 11, 11, 13, 3, 20, 16, 11, 11, 11, 11, 13]):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A3"


def write_output_workbook(
    rtgs_summary: dict,
    delay_summary: dict,
    receipt_made_summary: dict,
    cash_mode_validation_summary: pd.DataFrame,
    rcpt_cxn: pd.DataFrame,
    extra_tabs: dict[str, pd.DataFrame] | None = None,
) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Receipt made summary")
    _write_receipt_made_summary_sheet(ws, receipt_made_summary)

    ws = wb.create_sheet("RTGS Summary")
    _write_zone_tat_sheet(ws, "RTGS Summary", rtgs_summary)

    ws = wb.create_sheet("Cash Mode Validat Summary")
    _write_df(ws, cash_mode_validation_summary)

    ws = wb.create_sheet("Delay in RCPTING Summary")
    _write_zone_tat_sheet(ws, "Delay in RCPTING Summary", delay_summary)

    ws = wb.create_sheet("RCPT CXN")
    _write_df(ws, rcpt_cxn)

    for name, df in (extra_tabs or {}).items():
        ws = wb.create_sheet(name[:31])
        _write_df(ws, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
